#! /usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
# from openpyxl.cell import get_column_letter, column_index_from_string

# 用openpyxl模块打开Excel文档
wb = openpyxl.load_workbook('test.xlsx')
# print(type(wb))
# print(wb.get_sheet_names())
# 从工作簿中取得工作表
sheet1 = wb.get_sheet_by_name('Sheet1')
# print(sheet)
# print(type(sheet))
# print(sheet.title)
# anotherSheet = wb.get_active_sheet()
# print(anotherSheet)

# 从表中取得单元格
# print(sheet['A1'])
# print(sheet['A1'].value)
# print(type(sheet['A1'].value))

# c = sheet['B1']
# print('Row: ' + str(c.row) + ', Column: ' + c.column + ' is: ' + c.value)
# print('Cell ' + c.coordinate + ' is ' + c.value)

# for i in range(1,8,2):
#   print(i, sheet.cell(row=i, column=2).value)

# 列字母和数字之间的转换
# print(get_column_letter(1))
# print(get_column_letter(2))
# print(get_column_letter(27))
# print(get_column_letter(900))
# print(column_index_from_string('A'))
# print(column_index_from_string('AA'))

# 从表中取得行和列
for rowOfCellObjects in sheet1['A1':'C3']:
  for cellObj in rowOfCellObjects:
    print(cellObj.coordinate, cellObj.value)
  print('--- END OF ROW ---')

